import openpyxl

def store_data_in_excel(roll_number, name, email, branch):
    # Load or create a workbook
    try:
        workbook = openpyxl.load_workbook('student_data.xlsx')
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    # Select the active sheet
    sheet = workbook.active

    # Check if the headers are present in the sheet, if not, add them
    if sheet.cell(row=1, column=1).value != 'Roll Number':
        sheet.cell(row=1, column=1, value='Roll Number')
        sheet.cell(row=1, column=2, value='Name')
        sheet.cell(row=1, column=3, value='Email')
        sheet.cell(row=1, column=4, value='Branch')

    # Find the next available row in the sheet
    next_row = sheet.max_row + 1

    # Write the data to the next available row
    sheet.cell(row=next_row, column=1, value=roll_number)
    sheet.cell(row=next_row, column=2, value=name)
    sheet.cell(row=next_row, column=3, value=email)
    sheet.cell(row=next_row, column=4, value=branch)

    # Save the workbook
    workbook.save('student_data.xlsx')

if __name__ == "__main__":
    while True:
        try:
            roll_number = int(input("Enter Roll Number (or enter 0 to stop): "))
            if roll_number == 0:
                break

            name = input("Enter Name: ")
            email = input("Enter Email: ")
            branch = input("Enter Branch: ")

            # Store data in Excel
            store_data_in_excel(roll_number, name, email, branch)

            print("Data for Roll Number {} stored successfully in the Excel sheet.".format(roll_number))
        except ValueError:
            print("Invalid input. Roll Number should be a number.")

    print("Data entry completed. Exiting...")
